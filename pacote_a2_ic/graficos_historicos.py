import openpyxl as opx
from openpyxl import load_workbook
import plotly.graph_objs as go
import qrcode
import pandas as pd
import os
from datetime import datetime


def graficos_acoes(dicionario_historico_acoes):
    wb = opx.load_workbook(filename='BANCO_GLLYT.xlsx')
    contador = 0
    for acao in dicionario_historico_acoes.keys():
        if contador == 3:
            break
        else:
            historico = dicionario_historico_acoes.get(acao)
            try:
                fig = go.Figure()
                fig.add_trace(go.Candlestick(x=historico.index,open = historico['Open'], high=historico['High'], low=historico['Low'], close=historico['Close']))
                fig.update_layout(title = f'Variação de {acao} ao longo do tempo')
                fig.write_image(f'grafico_acao{contador}.jpg')
                img = opx.drawing.image.Image(f'grafico_acao{contador}.jpg')
                ws = wb.create_sheet(title=f"Gráficos_ações{contador}")
                ws.add_image(img, "A1")
            except:
                historico = historico
        contador+=1
    wb.save(filename='BANCO_GLLYT.xlsx' )







def graficos_moedas(dicionario_historico_moedas):
    wb = opx.load_workbook(filename='BANCO_GLLYT.xlsx')
    contador = 0
    for moeda in dicionario_historico_moedas.keys():
        if contador == 3:
            break
        else:
            historico = dicionario_historico_moedas.get(moeda)
            try:
                fig = go.Figure()
                fig.add_trace(go.Candlestick(x=historico.index,open = historico['Open'], high=historico['High'], low=historico['Low'], close=historico['Close']))
                fig.update_layout(title = f'Variação de {moeda} ao longo do tempo')
                fig.write_image(f'grafico_moeda{contador}.jpg')
                img = opx.drawing.image.Image(f'grafico_moeda{contador}.jpg')
                ws = wb.create_sheet(title=f"Gráficos_moedas{contador}")
                ws.add_image(img, "A1")
            except:
                historico = historico
        contador+=1
    wb.save(filename='BANCO_GLLYT.xlsx' )
    

