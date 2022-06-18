#Importe as bibliotecas necessárias
import openpyxl as opx
from openpyxl import load_workbook
from openpyxl.drawing.image import Image 
import qrcode
import pandas as pd
from openpyxl.utils import get_column_letter
from glob import glob


def dados(ws, dicio_ativo, dicio_cota, coluna):     # Adiciona os dados dos dicionários na planilha

    linha = 2
    for nome_ativo, quant_ativo in dicio_ativo.items():         # Desempacota os dados do dicionário do ativo
        if type(dicio_cota[nome_ativo]) == float:
            cotacao_ativo = quant_ativo * dicio_cota[nome_ativo]    # Calcula o valor total pareado ao real
            ws.cell(column=coluna, row=linha, value=nome_ativo)
            ws.cell(column=coluna+1, row=linha, value=quant_ativo)
            ws.cell(column=coluna+2, row=linha, value=cotacao_ativo).number_format = 'R$#,##0.00'
            # Adiciona os dados do ativo nas celulas especificadas
        else:
            ws.cell(column=coluna, row=linha, value=nome_ativo)
            ws.cell(column=coluna+1, row=linha, value=quant_ativo)
            ws.cell(column=coluna+2, row=linha, value='Não foi possível calcular').number_format = 'R$#,##0.00'
        linha += 1
    return ws


def redimensionar(ws):  # Redimensiona as colunas da planilha

    dimensoes = {}
    for linha in ws.rows:
        for celula in linha:
            if celula.value:
                dimensoes[celula.column_letter] = max((dimensoes.get(celula.column_letter, 0), len(str(celula.value))))
                # Armazena o tamanho do maior valor das celulas para cada coluna

    for coluna, valor in dimensoes.items():
        ws.column_dimensions[coluna].width = valor + 10
        # Usa os valores armazenados para redimensionar cada coluna
    return ws





def planilha(dicio_moedas, dicio_acoes, moedas_cota, acoes_cota):   # Cria uma planilha com os dados da carteira

    wb = opx.Workbook()    # Cria uma pasta de trabalho

    ws1 = wb.active             # Recebe a primeira planilha ativa
    ws1.title = "Carteira"      # Nomeia a planilha

    ws1.append(["Moedas", "Quantidade", "Valor total", "", "Ações", "Quantidade", "Valor total"])
    # Adiciona a lista como linha na planilha

    ws1 = dados(ws1, dicio_moedas, moedas_cota, 1)  # Adiciona os dados das moedas na planilha
    ws1 = dados(ws1, dicio_acoes, acoes_cota, 5)    # Adiciona os dados das ações na planilha

    ws1 = redimensionar(ws1)    # Redimensiona as colunas da planilha

    wb.save(filename='BANCO_GLLYT.xlsx')     # Salva a pasta de trabalho

    



#Função que lê os dicionários de moedas e ações e suas cotações e retorna o total (em reais) da carteira.
def valor_total(moedas, acoes, moedas_cota, acoes_cota):
    soma=0
    for moeda, quant_moeda in moedas.items():
        try:
            soma  += quant_moeda * moedas_cota[moeda]
        except:
            soma=soma
        
    for acao, quant_acao in acoes.items():
        try:
            soma += quant_acao * acoes_cota[acao]
        except:
            soma = soma
    return soma


#Função que insere um qrcode com o valor total da carteira na planilha excel da mesma.
def qr(soma):
    wb = opx.load_workbook(filename='BANCO_GLLYT.xlsx' )
    ws = wb.create_sheet(title="QrCode")
    qr = qrcode.QRCode(version = 1, error_correction = qrcode.constants.ERROR_CORRECT_H,box_size = 10, border = 4)
    qr.add_data(soma)
    img=qr.make_image()
    img.save('total.png')
    img = opx.drawing.image.Image('total.png')
    ws.add_image(img, "A1")
    wb.save(filename='BANCO_GLLYT.xlsx')
