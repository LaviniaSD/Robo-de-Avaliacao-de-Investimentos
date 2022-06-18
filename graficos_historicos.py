import openpyxl as opx
from openpyxl import load_workbook
import plotly.graph_objs as go
import qrcode
import pandas as pd
import os
from datetime import datetime


def graficos(dicionario_cotacoes_moedas, dicionario_preco_acoes):
    wb = opx.load_workbook(filename= 'Carteira.xlsx' )
    contador = 0
    for moeda in dicionario_cotacoes_moedas.keys():
        print(moeda)
        if contador == 3:
            break
        else:
            historico = dicionario_cotacoes_moedas.get(moeda)
            print(historico)
            try:
                fig = go.Figure()
                fig.add_trace(go.Candlestick(x=historico.index,open = historico['Open'], high=historico['High'], low=historico['Low'], close=historico['Close']))
                fig.update_layout(title = f'Variação de {moeda} ao longo do tempo')
                fig.write_image(f'grafico_moeda{contador}.jpg')
                img = opx.drawing.image.Image(f'grafico_moeda{contador}.jpg')
                ws = wb.create_sheet(title=f"Gráficos_{contador}")
                ws.add_image(img, "A1")
            except:
                historico = historico
        contador+=1
    wb.save(filename='Carteira.xlsx' )
    
'''                
    for historico in dicionario_preco_acoes.values():
        fig = go.Figure()
        fig.add_trace(go.Candlestick(x=historico.index,open = historico['Open'], high=historico['High'], low=historico['Low'], close=historico['Close'], name = 'market data'))
        fig.update_layout(title = ' Google share price', yaxis_title = 'Stock Price (USD)')
        fig.write_image('graficos_acoes.jpg')              
'''
