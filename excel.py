import openpyxl
from openpyxl.utils import get_column_letter
from glob import glob


def dados(ws, dicio_ativo, dicio_cota, coluna):     # Adiciona os dados dos dicionários na planilha

    linha = 2
    for nome_ativo, quant_ativo in dicio_ativo.items():         # Desempacota os dados do dicionário do ativo

        cotacao_ativo = quant_ativo * dicio_cota[nome_ativo]    # Calcula a cotação

        ws.cell(column=coluna, row=linha, value=nome_ativo)
        ws.cell(column=coluna+1, row=linha, value=quant_ativo)
        ws.cell(column=coluna+2, row=linha, value=cotacao_ativo).number_format = 'R$#,##0.00'
        # Adiciona os dados do ativo nas celulas especificadas

        linha += 1
    return ws


def redimensionar(ws):  # Redimensiona as colunas da planilha

    dimensoes = {}
    for linha in ws.rows:
        for celula in linha:
            if celula.value:
                dimensoes[celula.column_letter] = max((dimensoes.get(celula.column_letter, 0), len(str(celula.value))))
                # Armazena o tamanho do maior valor das celulas para cada coluna

    for coluna, valor in dimensoes.items():
        ws.column_dimensions[coluna].width = valor + 10
        # Usa os valores armazenados para redimensionar cada coluna
    return ws


def nome_planilha():    # Gera o nome da pasta de trabalho com base nos outros arquivos presentes no diretorio atual

    lista_arquivos = glob("Carteira*.xlsx")
    # Cria uma lista com os nomes dos arquivos existentes no formato especificado

    numero = len(lista_arquivos) + 1

    if numero == 1:
        return "Carteira.xlsx"
        # Retorna o nome do arquivo
    else:
        return "Carteira (%d).xlsx" % numero
        # Retorna um novo nome do arquivo com indexador atualizado


def planilha(dicio_moedas, dicio_acoes, moedas_cota, acoes_cota):   # Cria uma planilha com os dados da carteira

    wb = openpyxl.Workbook()    # Cria uma pasta de trabalho

    ws1 = wb.active             # Recebe a primeira planilha ativa
    ws1.title = "Carteira"      # Nomeia a planilha

    ws1.append(["Moedas", "Quantidade", "Cotação", "", "Ações", "Quantidade", "Cotação"])
    # Adiciona a lista como linha na planilha

    ws1 = dados(ws1, dicio_moedas, moedas_cota, 1)  # Adiciona os dados das moedas na planilha
    ws1 = dados(ws1, dicio_acoes, acoes_cota, 5)    # Adiciona os dados das ações na planilha

    ws1 = redimensionar(ws1)    # Redimensiona as colunas da planilha

    dest_filename = nome_planilha()     # Recebe o nome da pasta de trabalho
    wb.save(filename=dest_filename)     # Salva a pasta de trabalho
